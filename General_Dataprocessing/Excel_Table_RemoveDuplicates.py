"""
This script reads in a specified table format (Excel worksheet from an Excel
workbook or plain csv file) as pd DataFrame and removes all duplicated rows via
pd drop_duplicates()
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def remove_duplicated_rows(path, type, worksheet_name="worksheet_name"):
    """ Remove all duplicated rows from a table contained in an Excel worksheet.
    :param path: Pathlib path to Excel file
    :param type: "Excel" or "csv" file?
    :param worksheet_name: The name of the worksheet containing the table
    :returns The function will save to a new file called "No_Duplicates.xlsx"
    in the same folder as the original file.
    """
    path = Path(path)
    if type == "Excel":
        wb = load_workbook(path)
        if worksheet_name not in wb.sheetnames:
            raise Exception("Worksheet not contained in Excel file!")
        table = pd.read_excel(path, sheet_name=worksheet_name)
    elif type == "csv":
        table = pd.read_csv(path)
    print("Table has " + str(table.shape[0]) + " rows.")
    rows_before = table.shape[0]
    table.drop_duplicates(inplace=True)
    rows_after = table.shape[0]
    delta = rows_before - rows_after
    if delta >0:
        print(str(delta) + " duplicated rows were dropped.")
        print("Table now has " + str(rows_after) + " rows.")
        table.to_excel(path.parent / "No_Duplicates.xlsx")

    if delta == 0:
        print("No duplicated rows detected")
        print("Table has " + str(rows_after) + " rows.")

if __name__ == "__main__":
    path = r"D:\OneDrive\Dokumente\Timetracking_All_Merged.csv"
    remove_duplicated_rows(path, 'csv')